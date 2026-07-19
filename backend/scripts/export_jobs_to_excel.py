#!/usr/bin/env python3
"""Export active scraped jobs from the Code Quest DB or API to a detailed Excel file."""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

# Allow running from repo root or backend/
BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

PROFILE_LABELS = {
    "internship_india": "Internships",
    "fresher_india": "Fresher Jobs",
    "entry_level_india": "Entry Level",
    "experienced_manual_india": "1+ Experience (manual)",
}
PROFILE_TIERS = {
    "internship_india": "intern_fresher",
    "fresher_india": "intern_fresher",
    "entry_level_india": "entry",
    "experienced_manual_india": "experienced",
}

ROLE_FAMILIES: list[tuple[str, tuple[str, ...]]] = [
    ("Data Science & ML", ("data scientist", "machine learning", "ml engineer", "ai engineer", "deep learning", "nlp", "computer vision")),
    ("Data Analytics", ("data analyst", "business analyst", "bi analyst", "analytics analyst", "reporting analyst")),
    ("Data Engineering", ("data engineer", "etl developer", "data pipeline", "big data engineer")),
    ("Software Engineering", ("software engineer", "software developer", "developer", "programmer", "sde", "full stack", "fullstack", "frontend", "backend", "web developer", "application developer")),
    ("DevOps & Cloud", ("devops", "site reliability", "sre", "cloud engineer", "platform engineer", "infrastructure engineer")),
    ("QA & Testing", ("qa engineer", "quality assurance", "test engineer", "sdet", "automation tester")),
    ("Product & Management", ("product manager", "project manager", "scrum master", "program manager", "business development")),
    ("Design", ("ux designer", "ui designer", "product designer", "graphic designer")),
    ("Database & DBA", ("database administrator", "dba", "sql developer", "oracle developer")),
    ("Cybersecurity", ("security analyst", "cyber security", "information security", "soc analyst")),
    ("Sales & Customer", ("sales executive", "inside sales", "customer care", "business development executive", "account manager")),
]

SKILL_KEYWORDS: tuple[str, ...] = (
    "Python", "Java", "JavaScript", "TypeScript", "C++", "C#", "Go", "Rust", "Ruby", "PHP", "Swift", "Kotlin",
    "SQL", "PostgreSQL", "MySQL", "MongoDB", "Redis", "Oracle", "Snowflake", "BigQuery",
    "React", "Angular", "Vue", "Node.js", "Django", "Flask", "FastAPI", "Spring Boot", ".NET",
    "AWS", "Azure", "GCP", "Docker", "Kubernetes", "Terraform", "CI/CD", "Jenkins", "Git",
    "Power BI", "Tableau", "Excel", "Spark", "Hadoop", "Kafka", "Airflow", "dbt",
    "Machine Learning", "Deep Learning", "TensorFlow", "PyTorch", "Scikit-learn", "NLP",
    "REST API", "GraphQL", "Microservices", "Linux", "Agile", "Scrum",
    "Salesforce", "CRM", "SAP", "Figma",
)

EXPERIENCE_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("Internship", ("intern", "internship", "summer intern", "graduate intern")),
    ("Fresher / Entry Level", ("fresher", "entry level", "entry-level", "graduate trainee", "trainee", "junior", "0-1 year", "0 to 1")),
    ("1+ years", ("1 year", "1+ year", "1-2 year", "1 to 2", "minimum 1 year")),
    ("2+ years", ("2 year", "2+ year", "2-3 year", "2 to 3")),
    ("3+ years", ("3 year", "3+ year", "3-5 year", "3 to 5")),
    ("5+ years / Senior", ("5 year", "5+ year", "senior", "lead", "principal", "architect", "staff engineer")),
]

INGEST_EXPERIENCE = {
    "internship_india": "Internship",
    "fresher_india": "Fresher / Entry Level",
    "entry_level_india": "Entry Level",
    "experienced_manual_india": "1+ years",
}

PROD_DEFAULT_URL = "https://learn-coding-through-production.up.railway.app"


def _normalize_text(*parts: str | None) -> str:
    return " ".join(p.strip() for p in parts if p and p.strip()).lower()


def infer_role_family(title: str, description: str | None) -> str:
    haystack = _normalize_text(title, description)
    for family, keywords in ROLE_FAMILIES:
        if any(kw in haystack for kw in keywords):
            return family
    return "General / Other"


def infer_role(title: str) -> str:
    cleaned = " ".join(title.split()).strip()
    return cleaned or "Untitled role"


def infer_experience(title: str, description: str | None, ingest_profile: str | None) -> str:
    haystack = _normalize_text(title, description)
    for label, keywords in EXPERIENCE_RULES:
        if any(kw in haystack for kw in keywords):
            return label
    if ingest_profile and ingest_profile in INGEST_EXPERIENCE:
        return INGEST_EXPERIENCE[ingest_profile]
    return "Not specified"


def parse_description_skills(description: str | None) -> list[str]:
    if not description:
        return []
    match = re.search(r"skills?\s*:\s*([^\n]+)", description, flags=re.IGNORECASE)
    if not match:
        return []
    raw = match.group(1).strip()
    parts = re.split(r"[,;|]", raw)
    return [p.strip() for p in parts if p.strip()][:20]


def extract_skills(title: str, description: str | None, limit: int = 20) -> list[str]:
    explicit = parse_description_skills(description)
    if explicit:
        return explicit[:limit]

    haystack = _normalize_text(title, description)
    found: list[str] = []
    for skill in SKILL_KEYWORDS:
        pattern = re.escape(skill.lower())
        if re.search(rf"\b{pattern}\b", haystack) or skill.lower() in haystack:
            found.append(skill)
    return found[:limit]


def format_salary(min_val: float | None, max_val: float | None, currency: str | None) -> str:
    if min_val is None and max_val is None:
        return ""
    cur = f"{currency} " if currency else ""
    if min_val is not None and max_val is not None:
        return f"{cur}{min_val:g} - {max_val:g}"
    value = min_val if min_val is not None else max_val
    return f"{cur}{value:g}"


def clean_description(description: str | None) -> str:
    if not description:
        return ""
    text = re.sub(r"<br\s*/?>", " ", description, flags=re.IGNORECASE)
    return " ".join(text.split())


def fetch_jobs_from_api(base_url: str, page_size: int = 100) -> list[dict]:
    base_url = base_url.rstrip("/")
    all_jobs: list[dict] = []
    page = 1
    total = None

    while True:
        params = urlencode({"limit": page_size, "page": page})
        url = f"{base_url}/api/jobs?{params}"
        req = Request(url, headers={"Accept": "application/json"})
        try:
            with urlopen(req, timeout=120) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
        except (HTTPError, URLError) as exc:
            raise SystemExit(f"Failed to fetch jobs from {url}: {exc}") from exc

        jobs = payload.get("jobs") or []
        if total is None:
            total = int(payload.get("total") or 0)
            print(f"API reports {total} active jobs", file=sys.stderr)

        if not jobs:
            break

        all_jobs.extend(jobs)
        print(f"Fetched page {page}: {len(jobs)} jobs ({len(all_jobs)}/{total})", file=sys.stderr)

        if len(all_jobs) >= total or len(jobs) < page_size:
            break

        page += 1
        time.sleep(0.15)

    return all_jobs


def fetch_active_jobs_from_db() -> list[dict]:
    from sqlalchemy.orm import Session

    from app.core.database import SessionLocal
    from app.models.models import ScrapedJob
    from app.services.job_store import scraped_job_to_dict

    db: Session = SessionLocal()
    try:
        rows = (
            db.query(ScrapedJob)
            .filter(ScrapedJob.link_status == "active")
            .order_by(ScrapedJob.created_at.desc())
            .all()
        )
        return [scraped_job_to_dict(row) for row in rows]
    finally:
        db.close()


def api_job_to_row(job: dict, index: int) -> dict[str, object]:
    title = str(job.get("title") or "")
    description = job.get("description")
    ingest_profile = job.get("ingestProfile")
    skills = extract_skills(title, description)
    profile_key = str(ingest_profile or "")
    salary_min = job.get("salaryMin")
    salary_max = job.get("salaryMax")

    return {
        "S.No": index,
        "Code Quest Job ID": job.get("jobId") or "",
        "Internal ID": job.get("id") or "",
        "Role": infer_role(title),
        "Role Family": infer_role_family(title, description),
        "Skills": ", ".join(skills),
        "Experience Level": infer_experience(title, description, ingest_profile),
        "Scrape Profile": PROFILE_LABELS.get(profile_key, profile_key or "Unknown"),
        "Experience Tier": PROFILE_TIERS.get(profile_key, ""),
        "Company": job.get("company") or "",
        "Location": job.get("location") or "",
        "Source": job.get("source") or "",
        "Job Type": job.get("jobType") or "",
        "Date Posted": job.get("datePosted") or "",
        "Salary Range": format_salary(salary_min, salary_max, job.get("currency")),
        "Salary Min": salary_min if salary_min is not None else "",
        "Salary Max": salary_max if salary_max is not None else "",
        "Currency": job.get("currency") or "",
        "Job Link": job.get("jobUrl") or "",
        "Apply Link": job.get("applyUrl") or job.get("jobUrl") or "",
        "Link Status": job.get("linkStatus") or "active",
        "Ingested At (UTC)": job.get("createdAt") or "",
        "Ingested At (IST)": job.get("createdAtIST") or "",
        "Description": clean_description(description),
    }


def export_to_excel(rows: list[dict[str, object]], output_path: Path) -> None:
    if not rows:
        raise SystemExit("No active jobs found.")

    headers = list(rows[0].keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "Active Jobs"
    ws.freeze_panes = "A2"

    header_font = Font(bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    link_font = Font(color="0563C1", underline="single")
    link_columns = {"Job Link", "Apply Link"}

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if header in link_columns and value:
                cell.hyperlink = str(value)
                cell.font = link_font
            if header == "Description":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        if header == "Description":
            ws.column_dimensions[letter].width = 60
        elif header in link_columns:
            ws.column_dimensions[letter].width = 42
        else:
            max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 36)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export all active scraped jobs to Excel")
    parser.add_argument(
        "--source",
        choices=("api", "db"),
        default="api",
        help="Fetch from production/local API (api) or local SQLite DB (db)",
    )
    parser.add_argument(
        "--base-url",
        default=PROD_DEFAULT_URL,
        help=f"Jobs API base URL when --source=api (default: {PROD_DEFAULT_URL})",
    )
    parser.add_argument("--page-size", type=int, default=100, help="API page size (max 100)")
    parser.add_argument(
        "--output",
        default="exports/codequest_active_jobs_detailed.xlsx",
        help="Output .xlsx path",
    )
    args = parser.parse_args()

    if args.source == "api":
        jobs = fetch_jobs_from_api(args.base_url, page_size=min(args.page_size, 100))
    else:
        jobs = fetch_active_jobs_from_db()

    jobs = [j for j in jobs if j.get("jobUrl")]
    rows = [api_job_to_row(job, idx) for idx, job in enumerate(jobs, start=1)]

    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = (BACKEND_DIR.parent / output_path).resolve()

    export_to_excel(rows, output_path)
    print(f"Exported {len(rows)} active jobs to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
