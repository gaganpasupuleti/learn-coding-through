"""Parse and validate job bulk import Excel files (fixed schema)."""

from __future__ import annotations

import io
import json
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.models import LearningBatch

# Columns admins may include (row 1). Only title, company_name, location are required headers.
KNOWN_COLUMNS = (
    "title",
    "company_name",
    "location",
    "employment_type",
    "description",
    "eligible_batch_id",
    "eligible_batch_name",
)

REQUIRED_HEADERS = ("title", "company_name", "location")

# First-row header aliases (after lowercasing; spaces → underscores). Helps spreadsheets that use friendly names.
_EXCEL_HEADER_ALIASES: dict[str, str] = {
    "job_title": "title",
    "position": "title",
    "role": "title",
    "company": "company_name",
    "employer": "company_name",
    "organization": "company_name",
    "org": "company_name",
    "job_location": "location",
    "work_location": "location",
    "office_location": "location",
    "city": "location",
}


def _normalize_excel_header(cell: Any) -> str:
    key = _cell_str(cell).lower().replace(" ", "_").replace("-", "_")
    while "__" in key:
        key = key.replace("__", "_")
    return _EXCEL_HEADER_ALIASES.get(key, key)


def build_job_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "jobs"
    ws.append(list(KNOWN_COLUMNS))
    ws.append(
        [
            "Backend Engineer",
            "Acme Inc",
            "Bengaluru",
            "Full-time",
            "REST APIs, SQL, teamwork",
            "",
            "",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_job_import_xlsx(data: bytes, db: Session) -> tuple[list[dict[str, Any]], list[tuple[int, str]], int]:
    """
    Returns (valid_row_dicts, errors, skipped_empty_rows).
    If errors is non-empty, the caller must not insert any rows (strict file-wide validation).
    """
    errors: list[tuple[int, str]] = []
    skipped = 0
    pending: list[dict[str, Any]] = []
    wb = None
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return [], [(0, f"Invalid Excel file: {exc}")], 0

    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return [], [(1, "Sheet is empty")], 0

        header_map: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            key = _normalize_excel_header(cell)
            if key in KNOWN_COLUMNS:
                header_map[key] = idx

        for required in REQUIRED_HEADERS:
            if required not in header_map:
                return [], [(1, f"Missing required column: {required}")], 0

        def col(row: tuple[Any, ...], name: str) -> str:
            idx = header_map.get(name)
            if idx is None or idx >= len(row):
                return ""
            return _cell_str(row[idx])

        def row_is_blank(row: tuple[Any, ...]) -> bool:
            for idx in header_map.values():
                if idx < len(row) and _cell_str(row[idx]):
                    return False
            return True

        excel_row = 1
        for data_row in rows_iter:
            excel_row += 1
            if row_is_blank(data_row):
                skipped += 1
                continue

            row_err: list[str] = []
            title = col(data_row, "title")
            company = col(data_row, "company_name")
            location = col(data_row, "location")
            emp = col(data_row, "employment_type") or "Full-time"
            desc_raw = col(data_row, "description")
            id_raw = col(data_row, "eligible_batch_id")
            batch_name = col(data_row, "eligible_batch_name")

            if len(title) < 2:
                row_err.append("title must be at least 2 characters")
            if len(company) < 2:
                row_err.append("company_name must be at least 2 characters")
            if len(location) < 2:
                row_err.append("location must be at least 2 characters")
            if len(emp) < 2:
                row_err.append("employment_type invalid")

            desc: str | None = desc_raw or None
            if desc and len(desc) > 2000:
                row_err.append("description exceeds 2000 characters")

            batch_id: int | None = None
            if id_raw:
                if not re.fullmatch(r"-?\d+", id_raw):
                    row_err.append("eligible_batch_id must be an integer")
                else:
                    batch_id = int(id_raw)
                    b = db.query(LearningBatch).filter(LearningBatch.id == batch_id).first()
                    if not b:
                        row_err.append(f"eligible_batch_id {batch_id} not found")

            resolved_name_id: int | None = None
            if batch_name:
                bn = batch_name.strip()
                b = (
                    db.query(LearningBatch)
                    .filter(func.lower(LearningBatch.name) == bn.lower())
                    .first()
                )
                if not b:
                    row_err.append(f'eligible_batch_name "{batch_name}" not found')
                else:
                    resolved_name_id = b.id

            if batch_id is not None and resolved_name_id is not None and batch_id != resolved_name_id:
                row_err.append("eligible_batch_id and eligible_batch_name refer to different batches")

            if row_err:
                for msg in row_err:
                    errors.append((excel_row, msg))
                continue

            final_batch_id = batch_id if batch_id is not None else resolved_name_id

            pending.append(
                {
                    "title": title[:200],
                    "company_name": company[:180],
                    "location": location[:120],
                    "employment_type": emp[:80],
                    "description": desc[:2000] if desc else None,
                    "eligible_batch_id": final_batch_id,
                }
            )

        return pending, errors, skipped
    finally:
        if wb is not None:
            wb.close()


def _linkedin_scalar_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        parts = [_linkedin_scalar_str(v) for v in value]
        return ", ".join(p for p in parts if p).strip()
    if isinstance(value, dict):
        for sub in (
            "name",
            "fullName",
            "formatted",
            "defaultLocalizedName",
            "cityName",
            "countryName",
            "text",
        ):
            if sub in value:
                inner = _linkedin_scalar_str(value[sub])
                if inner:
                    return inner
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _linkedin_field(item: dict[str, Any], *keys: str) -> str:
    for key in keys:
        if key not in item:
            continue
        s = _linkedin_scalar_str(item[key])
        if s:
            return s
    return ""


def _linkedin_extract_rows(payload: Any) -> list[Any] | None:
    """Accept raw array or common wrapper objects (Apify, dataset exports, etc.)."""
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("data", "jobs", "items", "results", "records", "rows", "value"):
            inner = payload.get(key)
            if isinstance(inner, list):
                return inner
    return None


def parse_linkedin_jobs_json(data: bytes) -> tuple[list[dict[str, Any]], list[tuple[int, str]], int]:
    """
    Parse LinkedIn-style job scraper JSON: a UTF-8 JSON array of objects with at least
    title, company, location (and optional jobUrl, seniorityLevel, etc.).

    Returns (row_dicts ready for JobPost insert extras, errors with 1-based row index, skipped_count).
    Duplicate jobUrl values within the file are skipped (counted in skipped).
    Invalid objects are reported in errors; valid rows are still returned (partial import).
    """
    errors: list[tuple[int, str]] = []
    skipped = 0
    pending: list[dict[str, Any]] = []
    try:
        text = data.decode("utf-8-sig")
        payload = json.loads(text)
    except UnicodeDecodeError:
        return [], [(0, "File must be UTF-8 JSON")], 0
    except json.JSONDecodeError as exc:
        return [], [(0, f"Invalid JSON: {exc}")], 0

    rows_list = _linkedin_extract_rows(payload)
    if rows_list is None:
        return (
            [],
            [
                (
                    0,
                    "JSON must be an array of jobs, or an object with a "
                    "'data' / 'jobs' / 'items' / 'results' / 'records' array field.",
                )
            ],
            0,
        )

    seen_urls: set[str] = set()

    for idx, raw in enumerate(rows_list):
        row_num = idx + 1
        if not isinstance(raw, dict):
            errors.append((row_num, "Each entry must be a JSON object"))
            continue

        title = _linkedin_field(
            raw,
            "title",
            "jobTitle",
            "position",
            "job_title",
            "name",
            "role",
        )
        company = _linkedin_field(
            raw,
            "company",
            "company_name",
            "companyName",
            "employerName",
            "employer",
            "organizationName",
            "organization",
        )
        location = _linkedin_field(
            raw,
            "location",
            "jobLocation",
            "formattedLocation",
            "locations",
            "job_location",
        )

        if not title and not company and not location:
            skipped += 1
            continue

        row_err: list[str] = []
        if len(title) < 2:
            row_err.append("title must be at least 2 characters")
        if len(company) < 2:
            row_err.append("company must be at least 2 characters")
        if len(location) < 2:
            row_err.append("location must be at least 2 characters")

        if row_err:
            for msg in row_err:
                errors.append((row_num, msg))
            continue

        job_url = _linkedin_field(
            raw,
            "jobUrl",
            "job_url",
            "link",
            "url",
            "jobLink",
            "applyUrl",
            "externalUrl",
            "applicationUrl",
        )
        if job_url:
            if job_url in seen_urls:
                skipped += 1
                continue
            seen_urls.add(job_url)

        emp = _linkedin_field(raw, "employmentType", "employment_type") or "Full-time"

        lines: list[str] = []
        for label, key in (
            ("Seniority", "seniorityLevel"),
            ("Function", "jobFunction"),
            ("Industries", "industries"),
            ("Applicants", "applicantCount"),
            ("Posted", "postedAt"),
        ):
            val = _linkedin_field(raw, key)
            if val:
                lines.append(f"{label}: {val}")
        desc: str | None = "\n".join(lines) if lines else None
        if desc and len(desc) > 2000:
            desc = desc[:1997] + "..."

        metadata: dict[str, Any] = {"source": "linkedin_scrape"}
        for meta_key, json_key in (
            ("seniorityLevel", "seniorityLevel"),
            ("employmentType", "employmentType"),
            ("jobFunction", "jobFunction"),
            ("industries", "industries"),
            ("applicantCount", "applicantCount"),
            ("postedAt", "postedAt"),
            ("jobUrl", "jobUrl"),
            ("companyUrl", "companyUrl"),
        ):
            v = _linkedin_field(raw, json_key)
            if v:
                metadata[meta_key] = v

        if job_url:
            metadata["jobUrl"] = job_url

        ext_url = job_url[:2048] if job_url else None

        pending.append(
            {
                "title": title[:200],
                "company_name": company[:180],
                "location": location[:120],
                "employment_type": emp[:80],
                "description": desc,
                "eligible_batch_id": None,
                "external_apply_url": ext_url,
                "listing_metadata": metadata,
            }
        )

    return pending, errors, skipped
