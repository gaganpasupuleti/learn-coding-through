"""
export_jobs_excel.py
--------------------
Pull the latest N jobs from the Code Quest DB and write a color-coded Excel file.

Usage (run from repo root):
    pip install openpyxl python-dotenv
    python scripts/export_jobs_excel.py

    # Custom limit or output path:
    python scripts/export_jobs_excel.py --limit 200 --out jobs_export.xlsx

    # Only active jobs (default). Include expired/link-failed too:
    python scripts/export_jobs_excel.py --all-statuses

Reads DATABASE_URL from backend/.env (falls back to the default SQLite at
backend/career_portal.db if the env file or variable is missing).

Requires: openpyxl, python-dotenv
For PostgreSQL: also requires psycopg2-binary  (pip install psycopg2-binary)
"""

from __future__ import annotations

import argparse
import os
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Dependency check
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is not installed. Run:  pip install openpyxl")

try:
    from dotenv import dotenv_values
except ImportError:
    sys.exit("python-dotenv is not installed. Run:  pip install python-dotenv")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parents[1]
BACKEND_DIR = REPO_ROOT / "backend"
ENV_FILE = BACKEND_DIR / ".env"
DEFAULT_SQLITE = BACKEND_DIR / "career_portal.db"

# Row colours per source (hex fill, no #)
SOURCE_COLORS: dict[str, str] = {
    "indeed":    "DDEEFF",   # light blue
    "linkedin":  "DDF0DD",   # light green
    "google":    "FFFACC",   # light yellow
    "naukri":    "FFE8CC",   # light orange
    "zip_recruiter": "EEE0FF",
    "glassdoor": "FFE0F0",
}
DEFAULT_ROW_COLOR = "F5F5F5"  # light grey for unknown sources

HEADER_FILL   = "1A2744"   # dark navy
HEADER_FONT   = "FFFFFF"   # white

# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def resolve_db_url() -> str:
    env = dotenv_values(ENV_FILE) if ENV_FILE.exists() else {}
    url = env.get("DATABASE_URL", "") or os.environ.get("DATABASE_URL", "")
    if url:
        # Normalize Railway shorthand
        if url.startswith("postgres://"):
            url = url.replace("postgres://", "postgresql://", 1)
        return url
    # Fall back to local SQLite
    return f"sqlite:///{DEFAULT_SQLITE}"


def fetch_jobs_sqlite(db_path: Path, limit: int, active_only: bool) -> list[dict]:
    if not db_path.exists():
        sys.exit(f"SQLite DB not found at {db_path}\nHave you run the backend at least once?")
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    status_clause = "WHERE link_status = 'active'" if active_only else ""
    rows = conn.execute(
        f"""
        SELECT id, job_id, source, title, company, location, job_type,
               date_posted, salary_min, salary_max, currency,
               job_url, apply_url, created_at, link_status, ingest_profile
        FROM scraped_jobs
        {status_clause}
        ORDER BY created_at DESC
        LIMIT ?
        """,
        (limit,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def fetch_jobs_postgres(url: str, limit: int, active_only: bool) -> list[dict]:
    try:
        import psycopg2
        import psycopg2.extras
    except ImportError:
        sys.exit("psycopg2-binary is not installed. Run:  pip install psycopg2-binary")

    status_clause = "WHERE link_status = 'active'" if active_only else ""
    conn = psycopg2.connect(url)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"""
        SELECT id, job_id, source, title, company, location, job_type,
               date_posted, salary_min, salary_max, currency,
               job_url, apply_url, created_at, link_status, ingest_profile
        FROM scraped_jobs
        {status_clause}
        ORDER BY created_at DESC
        LIMIT %s
        """,
        (limit,),
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def fetch_jobs(limit: int, active_only: bool) -> list[dict]:
    url = resolve_db_url()
    print(f"  DB: {url[:60]}{'...' if len(url) > 60 else ''}")

    if url.startswith("sqlite:///"):
        raw_path = url.replace("sqlite:///", "")
        # Resolve relative paths (e.g. "./career_portal.db") against backend dir
        db_path = Path(raw_path) if Path(raw_path).is_absolute() else BACKEND_DIR / raw_path.lstrip("./")
        return fetch_jobs_sqlite(db_path, limit, active_only)

    if "postgresql" in url or "postgres" in url:
        return fetch_jobs_postgres(url, limit, active_only)

    sys.exit(f"Unsupported DATABASE_URL scheme: {url[:40]}")


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def fmt_salary(row: dict) -> str:
    lo, hi, cur = row.get("salary_min"), row.get("salary_max"), row.get("currency") or ""
    if lo and hi:
        return f"{cur} {int(lo):,} – {int(hi):,}".strip()
    if lo:
        return f"{cur} {int(lo):,}+".strip()
    if hi:
        return f"up to {cur} {int(hi):,}".strip()
    return ""


def fmt_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value[:10]
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)


IST_OFFSET = 19800   # Asia/Kolkata = UTC+5:30 in seconds


def fmt_ist(value) -> str:
    """Convert a UTC datetime (or ISO string) to 'D MMM YYYY, hh:mm AM/PM IST'.

    - Assumes naive datetimes are UTC (DB stores UTC without tzinfo).
    - If value already has tzinfo, converts without adding offset again.
    - If parsing fails, returns original value as string.
    """
    if value is None:
        return ""
    try:
        from datetime import timedelta, timezone
        if isinstance(value, str):
            # SQLite stores as "YYYY-MM-DD HH:MM:SS[.ffffff]"
            value = datetime.fromisoformat(value.replace(" ", "T").split(".")[0])
        if not isinstance(value, datetime):
            return str(value)
        # If already timezone-aware, convert directly; otherwise treat as UTC
        if value.tzinfo is not None:
            ist_dt = value.astimezone(timezone(timedelta(seconds=IST_OFFSET)))
        else:
            ist_dt = value + timedelta(seconds=IST_OFFSET)
        day = ist_dt.strftime("%d").lstrip("0") or "0"
        return f"{day} {ist_dt.strftime('%b %Y, %I:%M %p')} IST"
    except Exception:
        return str(value)


COLUMNS = [
    ("#",           6),
    ("Job ID",      18),
    ("Title",       42),
    ("Company",     24),
    ("Location",    22),
    ("Source",      12),
    ("Type",        12),
    ("Date Posted", 13),
    ("Salary",      22),
    ("Status",      12),
    ("Profile",     16),
    ("Apply Link",  80),
    ("Job URL",     80),
    ("Added",       26),
]


def build_excel(jobs: list[dict], out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"
    ws.freeze_panes = "A2"

    # Header
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color=HEADER_FONT, size=11)
    for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, job in enumerate(jobs, start=2):
        source = (job.get("source") or "").lower()
        fill_hex = SOURCE_COLORS.get(source, DEFAULT_ROW_COLOR)
        row_fill = PatternFill("solid", fgColor=fill_hex)

        apply_url = job.get("apply_url") or job.get("job_url") or ""
        job_url   = job.get("job_url") or ""

        values = [
            row_idx - 1,                                        # #
            job.get("job_id") or job.get("id") or "",          # Job ID (fallback to internal id)
            job.get("title") or "",                             # Title
            job.get("company") or "",                           # Company
            job.get("location") or "",                          # Location
            (job.get("source") or "").title(),                  # Source
            job.get("job_type") or "",                          # Type
            fmt_date(job.get("date_posted")),                   # Date Posted
            fmt_salary(job),                                    # Salary
            job.get("link_status") or "",                       # Status
            job.get("ingest_profile") or "",                    # Profile
            apply_url,                                          # Apply Link (col 12)
            job_url,                                            # Job URL   (col 13)
            fmt_ist(job.get("created_at")),                     # Added (IST)
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            # Style URL columns blue (cols 12 and 13)
            if col_idx in (12, 13) and value:
                cell.font = Font(color="0563C1", underline="single")

    # Column widths
    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Legend sheet
    ws_legend = wb.create_sheet("Legend")
    ws_legend["A1"] = "Source"
    ws_legend["B1"] = "Row Colour"
    ws_legend["A1"].font = Font(bold=True)
    ws_legend["B1"].font = Font(bold=True)
    legend_rows = [
        ("Indeed",       SOURCE_COLORS["indeed"]),
        ("LinkedIn",     SOURCE_COLORS["linkedin"]),
        ("Google",       SOURCE_COLORS["google"]),
        ("Naukri",       SOURCE_COLORS["naukri"]),
        ("ZipRecruiter", SOURCE_COLORS["zip_recruiter"]),
        ("Glassdoor",    SOURCE_COLORS["glassdoor"]),
        ("Other",        DEFAULT_ROW_COLOR),
    ]
    for i, (name, color) in enumerate(legend_rows, start=2):
        ws_legend.cell(row=i, column=1, value=name)
        c = ws_legend.cell(row=i, column=2, value="")
        c.fill = PatternFill("solid", fgColor=color)
    ws_legend.column_dimensions["A"].width = 16
    ws_legend.column_dimensions["B"].width = 18

    wb.save(str(out_path))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Export latest jobs from Code Quest DB to Excel.")
    parser.add_argument("--limit",        type=int, default=100,                    help="Number of jobs to export (default: 100)")
    parser.add_argument("--out",          type=str, default="jobs_export.xlsx",     help="Output file name (default: jobs_export.xlsx)")
    parser.add_argument("--all-statuses", action="store_true",                      help="Include expired and link-failed jobs (default: active only)")
    args = parser.parse_args()

    out_path = Path(args.out)
    active_only = not args.all_statuses

    print(f"\nCode Quest Jobs Export")
    print(f"  Limit    : {args.limit}")
    print(f"  Filter   : {'active only' if active_only else 'all statuses'}")
    print(f"  Output   : {out_path.resolve()}")

    jobs = fetch_jobs(args.limit, active_only)
    print(f"  Fetched  : {len(jobs)} jobs")

    if not jobs:
        print("\nNo jobs found. Check that jobs have been loaded into the DB.")
        sys.exit(0)

    # Source breakdown
    from collections import Counter
    breakdown = Counter((j.get("source") or "unknown").lower() for j in jobs)
    print("  Sources  :", ", ".join(f"{k}={v}" for k, v in sorted(breakdown.items())))

    build_excel(jobs, out_path)
    print(f"\nSaved: {out_path.resolve()}\n")


if __name__ == "__main__":
    main()
