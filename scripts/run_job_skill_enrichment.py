"""Rules-v1 auto enrichment: classify scraped jobs → role family, skills, commit + Excel."""

from __future__ import annotations

import csv
import io
import re
import sys
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path

BACKEND = Path(__file__).resolve().parents[1] / "backend"
sys.path.insert(0, str(BACKEND))

from app.core.database import SessionLocal
from app.data.job_role_seed import JOB_ROLES
from app.models.models import JobEnrichment, ScrapedJob
from app.services.job_enrichment_import import commit_job_enrichment_import, preview_job_enrichment_import

REPO = BACKEND.parent
OUT_CSV = REPO / "local-data" / "job-enrichment-auto-last-5d.csv"
OUT_XLSX = REPO / "india_jobs_enriched_last_5_days.xlsx"

PROFILES = (
    "internship_india",
    "fresher_india",
    "entry_level_india",
    "experienced_manual_india",
    "platform_crm_india",
    "ai_india",
)

# ponytail: ordered title rules — specific families before generic (see FRESHER_ROLE_FAMILY_CATALOG.md)
TITLE_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("ROLE_POWER_PLATFORM", ("power apps", "powerapps", "power automate", "power platform", "canvas app", "power fx", "model-driven app")),
    ("ROLE_POWERBI_ANALYST", ("power bi", "powerbi", " dax", "power-bi")),
    ("ROLE_SALESFORCE", ("salesforce", "sfdc", "apex developer", "salesforce developer", "salesforce admin", "lightning developer")),
    ("ROLE_DYNAMICS_CRM", ("dynamics 365", "dynamics crm", "microsoft dynamics", " d365", "crm developer")),
    ("ROLE_CYBER_SECURITY", ("cyber security", "cybersecurity", "soc analyst", "information security", "infosec", "security analyst", "security operations", "cyber defense", "soc ")),
    ("ROLE_DATA_ANALYST", ("data analyst", "reporting analyst", "bi analyst", "mis analyst", "analytics analyst")),
    ("ROLE_DATA_ENGINEER", ("data engineer", " etl", "pipeline engineer", "spark", "airflow")),
    ("ROLE_AGENTIC_AI", (
        "agentic ai", "ai agent", "ai agents", "multi-agent", "multi agent", "autonomous agent",
        "langgraph", "crewai", "autogen", "agent workflow", "agentic workflow", "agent developer",
    )),
    ("ROLE_GEN_AI", (
        "gen ai", "genai", "generative ai", "llm engineer", "llm developer", "large language model",
        "prompt engineer", "rag engineer", "rag developer", " chatgpt", " openai", " llm ", " ai engineer",
    )),
    ("ROLE_ML_AI", (
        "machine learning", "ml engineer", "data scientist", "data science engineer",
        "deep learning", "nlp engineer", "computer vision", "tensorflow", "pytorch", "scikit-learn",
    )),
    ("ROLE_JAVA_BACKEND", ("java developer", "java engineer", "spring boot", "springboot", " java ", "java backend")),
    ("ROLE_PYTHON_DEV", ("python developer", "python engineer", " django", " flask", " fastapi", " python ")),
    ("ROLE_FRONTEND_REACT", (
        "react developer", "react.js", "next.js", "nextjs", "frontend developer", "front end developer",
        "front-end", "frontend engineer", " ui developer", " react ", "javascript developer", "typescript developer",
    )),
    ("ROLE_QA_TESTING", ("qa engineer", "quality assurance", "software tester", "manual testing", "test engineer", "selenium")),
    ("ROLE_SERVICENOW", ("servicenow", "itsm developer", "itsm admin")),
    ("ROLE_IT_SUPPORT", ("it support", "helpdesk", "help desk", "desktop support", " l1 support", " l2 support")),
    ("ROLE_BUSINESS_ANALYST", ("business analyst", "functional analyst", "requirements analyst", " brd ", " uat ")),
    ("ROLE_FULLSTACK_WEB", ("full stack", "fullstack", "mern", "software engineer", "software developer", "sde", "web developer")),
]

ROLE_NAMES = {r["role_id"]: str(r["role_name"]) for r in JOB_ROLES}

SKILL_LEXICON: dict[str, tuple[str, str]] = {
    "java": ("lang", "java"),
    "python": ("lang", "python"),
    "javascript": ("lang", "javascript"),
    "typescript": ("lang", "typescript"),
    "sql": ("req", "sql"),
    "react": ("fw", "react"),
    "spring": ("fw", "spring-boot"),
    "django": ("fw", "django"),
    "flask": ("fw", "flask"),
    "fastapi": ("fw", "fastapi"),
    "node.js": ("fw", "nodejs"),
    "nodejs": ("fw", "nodejs"),
    "excel": ("tool", "excel"),
    "power bi": ("tool", "power-bi"),
    "powerbi": ("tool", "power-bi"),
    "tableau": ("tool", "tableau"),
    "git": ("tool", "git"),
    "jira": ("tool", "jira"),
    "selenium": ("tool", "selenium"),
    "postman": ("tool", "postman"),
    "aws": ("tool", "aws"),
    "azure": ("tool", "azure"),
    "mysql": ("db", "mysql"),
    "postgresql": ("db", "postgresql"),
    "mongodb": ("db", "mongodb"),
    "html": ("req", "html-css"),
    "css": ("req", "html-css"),
    "rest api": ("req", "rest-apis"),
    "api": ("req", "rest-apis"),
    "machine learning": ("req", "machine-learning"),
    "data analysis": ("req", "data-analysis"),
    "testing": ("req", "manual-testing"),
    "automation": ("req", "test-automation"),
    "salesforce": ("tool", "salesforce"),
    "apex": ("lang", "apex"),
    "crm": ("req", "crm"),
    "power apps": ("tool", "power-apps"),
    "powerapps": ("tool", "power-apps"),
    "power automate": ("tool", "power-automate"),
    "dynamics 365": ("tool", "dynamics-365"),
    "dynamics": ("tool", "dynamics-365"),
    "siem": ("tool", "siem"),
    "cybersecurity": ("req", "cybersecurity"),
    "cyber security": ("req", "cybersecurity"),
    "soc analyst": ("req", "soc-operations"),
    "next.js": ("fw", "nextjs"),
    "nextjs": ("fw", "nextjs"),
    "langchain": ("fw", "langchain"),
    "langgraph": ("fw", "langgraph"),
    "generative ai": ("req", "generative-ai"),
    "gen ai": ("req", "generative-ai"),
    "llm": ("req", "llm"),
    "rag": ("req", "rag"),
    "prompt engineering": ("req", "prompt-engineering"),
    "tensorflow": ("fw", "tensorflow"),
    "pytorch": ("fw", "pytorch"),
    "agentic": ("req", "agentic-ai"),
}

PREP_TOPICS = {
    "java": "java-basics",
    "python": "python-fundamentals",
    "javascript": "javascript-es6",
    "sql": "sql-queries",
    "react": "react-components",
    "html-css": "html-css",
    "excel": "excel-pivot",
    "power-bi": "power-bi-basics",
    "manual-testing": "testing-fundamentals",
    "rest-apis": "api-basics",
    "data-analysis": "data-visualization-basics",
    "machine-learning": "ml-basics",
    "generative-ai": "gen-ai-basics",
    "llm": "llm-fundamentals",
    "rag": "rag-basics",
    "agentic-ai": "agentic-ai-basics",
    "prompt-engineering": "prompt-engineering-basics",
    "cybersecurity": "security-fundamentals",
    "crm": "crm-basics",
    "power-apps": "power-platform-basics",
    "salesforce": "salesforce-basics",
}

YEARS_RE = re.compile(r"(\d+)\+?\s*(?:years?|yrs?)", re.I)
FRESHER_KW = ("fresher", "graduate", "trainee", "intern", "entry level", "entry-level", "0-1 year", "0 to 1")
ENTRY_KW = ("junior", "associate", "1 year", "1-2 year", "1 to 2", "1 yr")
EXP_KW = ("senior", "lead", "architect", "principal", "5+ year", "5 year", "4 year", "3-5 year")

CSV_COLS = [
    "job_id", "actual_role_id", "actual_role_name", "role_level_id", "experience_level",
    "job_live_status", "jd_fetch_status", "jd_summary", "required_skills", "good_to_have_skills",
    "tools", "programming_languages", "databases", "frameworks", "student_preparation_topics",
    "quiz_pack_id", "mapping_confidence", "manual_review_needed", "notes",
]


def _text(job: ScrapedJob) -> str:
    parts = [job.title or "", job.description or "", job.job_type or ""]
    return " ".join(parts).lower()


def _score_roles(text: str) -> list[tuple[str, float]]:
    scores: list[tuple[str, float]] = []
    for role_id, patterns in TITLE_RULES:
        hits = sum(1 for p in patterns if p in text)
        if hits:
            scores.append((role_id, min(0.95, 0.55 + hits * 0.15)))
    scores.sort(key=lambda x: -x[1])
    return scores


def _experience_band(text: str, profile: str | None) -> tuple[str, str]:
    m = YEARS_RE.search(text)
    if m:
        yrs = int(m.group(1))
        if yrs <= 1:
            return "fresher", "FRESHER"
        if yrs <= 2:
            return "entry", "ENTRY"
        return "experienced", "EXPERIENCED"

    if profile == "internship_india" or " intern" in f" {text}":
        return "fresher", "FRESHER"
    if profile == "fresher_india" or any(k in text for k in FRESHER_KW):
        return "fresher", "FRESHER"
    if profile == "entry_level_india" or any(k in text for k in ENTRY_KW):
        return "entry", "ENTRY"
    if profile == "experienced_manual_india" or any(k in text for k in EXP_KW):
        return "experienced", "EXPERIENCED"
    if profile == "entry_level_india":
        return "entry", "ENTRY"
    return "fresher", "FRESHER"


def _extract_skills(text: str) -> dict[str, list[str]]:
    buckets: dict[str, list[str]] = {
        "required_skills": [],
        "good_to_have_skills": [],
        "tools": [],
        "programming_languages": [],
        "databases": [],
        "frameworks": [],
    }
    counts: Counter[str] = Counter()
    for token, (bucket, slug) in SKILL_LEXICON.items():
        if token in text:
            counts[slug] += text.count(token)
            key = {
                "req": "required_skills",
                "good": "good_to_have_skills",
                "tool": "tools",
                "lang": "programming_languages",
                "db": "databases",
                "fw": "frameworks",
            }[bucket]
            if slug not in buckets[key]:
                buckets[key].append(slug)

    for slug, cnt in counts.items():
        if cnt >= 2 and slug not in buckets["required_skills"]:
            for k in ("good_to_have_skills", "tools", "programming_languages", "databases", "frameworks"):
                if slug in buckets[k]:
                    buckets[k].remove(slug)
            if slug not in buckets["required_skills"]:
                buckets["required_skills"].append(slug)

    if not buckets["required_skills"] and counts:
        top = counts.most_common(3)
        buckets["required_skills"] = [s for s, _ in top]

    return buckets


def _pipe(items: list[str]) -> str:
    return "|".join(items)


def _jd_status(desc: str | None) -> str:
    if not desc or len(desc.strip()) < 80:
        return "UNKNOWN"
    if len(desc.strip()) >= 400:
        return "FETCHED"
    return "PARTIAL"


def _summary(job: ScrapedJob, role_name: str) -> str:
    desc = (job.description or "").strip()
    if len(desc) >= 120:
        s = re.sub(r"\s+", " ", desc)[:480]
        return f"{role_name} role at {job.company or 'company'}. {s}…"
    return (
        f"Title/metadata classification: {job.title} at {job.company or 'company'}. "
        f"Scrape profile {job.ingest_profile or 'unknown'}."
    )


def classify_job(job: ScrapedJob) -> dict[str, str]:
    text = _text(job)
    scores = _score_roles(text)
    exp_level, exp_suffix = _experience_band(text, job.ingest_profile)

    if not scores:
        role_id = "ROLE_OTHER_REVIEW"
        confidence = 0.35
        needs_review = True
        notes = "No role family keyword match; mapped to ROLE_OTHER_REVIEW."
    else:
        role_id, confidence = scores[0]
        needs_review = False
        notes = f"rules-v1 title match → {role_id} (score={confidence:.2f})."
        if len(scores) > 1 and scores[1][1] >= scores[0][1] - 0.1:
            confidence = min(confidence, 0.55)
            needs_review = True
            notes += f" Ambiguous vs {scores[1][0]}."
        if role_id == "ROLE_FULLSTACK_WEB" and confidence < 0.75:
            needs_review = True
            notes += " Generic software engineer title."

    if exp_level == "fresher" and any(k in text for k in EXP_KW):
        needs_review = True
        notes += " Conflicting fresher title vs senior keywords."

    skills = _extract_skills(text)
    if not skills["required_skills"]:
        skills["required_skills"] = ["programming-fundamentals"]
        confidence = min(confidence, 0.65)
        needs_review = True
        notes += " No skills extracted; defaulted fundamentals."

    jd_status = _jd_status(job.description)
    if jd_status in ("UNKNOWN", "PARTIAL"):
        confidence = min(confidence, 0.69)
        needs_review = True

    if confidence < 0.70:
        needs_review = True

    if role_id == "ROLE_OTHER_REVIEW":
        needs_review = True

    role_level_id = f"{role_id}_{exp_suffix}"
    prep = []
    for slug in skills["required_skills"][:5]:
        if slug in PREP_TOPICS:
            prep.append(PREP_TOPICS[slug])

    return {
        "job_id": job.job_id or "",
        "actual_role_id": role_id,
        "actual_role_name": ROLE_NAMES.get(role_id, role_id),
        "role_level_id": role_level_id,
        "experience_level": exp_level,
        "job_live_status": "LIVE" if (job.link_status or "active") == "active" else "UNKNOWN",
        "jd_fetch_status": jd_status,
        "jd_summary": _summary(job, ROLE_NAMES.get(role_id, role_id)),
        "required_skills": _pipe(skills["required_skills"]),
        "good_to_have_skills": _pipe(skills["good_to_have_skills"]),
        "tools": _pipe(skills["tools"]),
        "programming_languages": _pipe(skills["programming_languages"]),
        "databases": _pipe(skills["databases"]),
        "frameworks": _pipe(skills["frameworks"]),
        "student_preparation_topics": _pipe(prep),
        "quiz_pack_id": "",
        "mapping_confidence": f"{confidence:.2f}",
        "manual_review_needed": "yes" if needs_review else "no",
        "notes": notes[:500],
    }


def fetch_jobs(db, days: int = 5) -> list[ScrapedJob]:
    cutoff = datetime.utcnow() - timedelta(days=days)
    return (
        db.query(ScrapedJob)
        .filter(
            ScrapedJob.link_status == "active",
            ScrapedJob.ingest_profile.in_(PROFILES),
            ScrapedJob.created_at >= cutoff,
            ScrapedJob.job_id.isnot(None),
        )
        .order_by(ScrapedJob.created_at.desc())
        .all()
    )


def fetch_all_active_jobs(db) -> list[ScrapedJob]:
    return (
        db.query(ScrapedJob)
        .filter(
            ScrapedJob.link_status == "active",
            ScrapedJob.job_id.isnot(None),
        )
        .order_by(ScrapedJob.created_at.desc())
        .all()
    )


def rows_to_csv_bytes(rows: list[dict[str, str]]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_COLS)
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def export_enriched_excel(db, out: Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    jobs = fetch_jobs(db, days=5)
    enrich_map = {e.job_id: e for e in db.query(JobEnrichment).all()}

    cols = [
        ("#", 5), ("Job ID", 18), ("Title", 40), ("Company", 22), ("Profile", 18),
        ("Role Family", 22), ("Role Level", 28), ("Experience", 12),
        ("Required Skills", 36), ("Tools", 24), ("Languages", 18), ("Frameworks", 18),
        ("Confidence", 10), ("Review?", 10), ("Apply Link", 60),
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enriched Jobs"
    ws.freeze_panes = "A2"
    hdr_fill = PatternFill("solid", fgColor="1A2744")
    hdr_font = Font(bold=True, color="FFFFFF")
    for i, (name, w) in enumerate(cols, 1):
        c = ws.cell(1, i, name)
        c.fill, c.font = hdr_fill, hdr_font
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, job in enumerate(jobs, 1):
        e = enrich_map.get(job.job_id)
        apply = job.apply_url or job.job_url or ""
        row = [
            idx, job.job_id, job.title, job.company, job.ingest_profile,
            e.actual_role_name if e else "",
            e.role_level_id if e else "",
            e.experience_level if e else "",
            "|".join(e.required_skills or []) if e and e.required_skills else "",
            "|".join(e.tools or []) if e and e.tools else "",
            "|".join(e.programming_languages or []) if e and e.programming_languages else "",
            "|".join(e.frameworks or []) if e and e.frameworks else "",
            float(e.mapping_confidence) if e and e.mapping_confidence is not None else "",
            "yes" if e and e.manual_review_needed else "",
            apply,
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(idx + 1, col, val)
            cell.alignment = Alignment(vertical="center")
            if col == 15 and val:
                cell.font = Font(color="0563C1", underline="single")
    wb.save(out)


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Rules-v1 job enrichment (classify → commit)")
    parser.add_argument("--all-active", action="store_true", help="Classify every active scraped job")
    parser.add_argument("--days", type=int, default=5, help="Window for default profile-scoped run")
    args = parser.parse_args()

    db = SessionLocal()
    try:
        if args.all_active:
            jobs = fetch_all_active_jobs(db)
            print(f"Classifying {len(jobs)} active jobs (all profiles, no date cutoff)")
        else:
            jobs = fetch_jobs(db, days=args.days)
            print(f"Classifying {len(jobs)} active jobs (last {args.days} days, target profiles)")
        rows = [classify_job(j) for j in jobs if j.job_id]
        rows = [r for r in rows if r["job_id"]]

        raw = rows_to_csv_bytes(rows)
        OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
        OUT_CSV.write_bytes(raw)
        print(f"Wrote CSV: {OUT_CSV}")

        preview = preview_job_enrichment_import(db, raw)
        print(
            f"Preview: total={preview.total_rows} valid={preview.valid_rows} "
            f"invalid={preview.invalid_rows} warnings={preview.warning_rows}"
        )
        if preview.invalid_rows:
            for err in preview.row_errors:
                if err.errors:
                    print(f"  row {err.row_number} {err.job_id}: {err.errors}")
            return 1

        result = commit_job_enrichment_import(db, raw)
        print(
            f"Committed: inserted={result.inserted_count} updated={result.updated_count} "
            f"skipped={result.skipped_count}"
        )

        by_role = Counter(r["actual_role_id"] for r in rows)
        print("Role breakdown:", dict(sorted(by_role.items(), key=lambda x: -x[1])))
        review = sum(1 for r in rows if r["manual_review_needed"] == "yes")
        print(f"Needs review: {review}/{len(rows)}")

        export_enriched_excel(db, OUT_XLSX)
        print(f"Excel: {OUT_XLSX.resolve()}")
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
